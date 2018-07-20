########################################################################################
'''
AUTHOR:			Yao Ji (jiyao94@126.com)
CREATED DATE:	2018/6/28
LAST UPDATE:	2018/7/20
DESCRIPTION: 	This tool is used to generate Arguments file accroding to Config file.
				Ti checks whether 'Config.xlsx' exists in the current directory. If not,
				it generates an empty one, otherwise it generates 'Arguments.xlsx'
				according to 'Config.xlsx'. User can specifie library, loop name, and
				start page in the config file.
'''
########################################################################################

import os, traceback
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.cell import Cell
from copy import copy

def CellCopy(ws, s_cell):
	t_cell = Cell(ws, value=s_cell.value)
	t_cell.font = copy(s_cell.font)
	t_cell.fill = copy(s_cell.fill)
	t_cell.alignment = copy(s_cell.alignment)
	t_cell.border = copy(s_cell.border)
	return t_cell

def Config(table, argsFileName='Arguments.xlsx'):
	wb = Workbook()
	ws = wb.active
	ws.title = 'Config'
	new_table = []
	for row in table:
		lst = []
		for col in row:
			if hasattr(col, 'value'):
				lst.append(col.value)
			else:
				lst.append(col)
		ws.append(lst)
		new_table.append(lst)
	table = new_table
	ws.column_dimensions['A'].width = 50.0
	ws.column_dimensions['B'].width = 50.0

	table.remove(table[0])
	for row in table:
		ws = wb.create_sheet(row[1])
		ws.merge_cells('A1:F1')
		wb_lib = load_workbook('./Library/' + row[0].replace('.txt', '.xlsx'))
		ws_lib = wb_lib.active
		table_lib = list(ws_lib.rows)
		for j in range(len(table_lib)):
			cell_list = []
			for k in range(len(table_lib[j])):
				cell_list.append(CellCopy(ws, table_lib[j][k]))
			ws.append(cell_list)
			if cell_list[0].value == 'INPUTS':
				ws.merge_cells('A{0}:C{0}'.format(j + 1))
			elif cell_list[0].value == 'OUTPUTS':
				ws.merge_cells('A{0}:D{0}'.format(j + 1))
			elif cell_list[0].value == 'PARAMETERS':
				ws.merge_cells('A{0}:F{0}'.format(j + 1))
			else:
				pass
			ws.column_dimensions['B'].width = 30.0
			ws.column_dimensions['C'].width = 15.0
			ws.column_dimensions['D'].width = 30.0

	wb.save(argsFileName)

#start main
if __name__ == '__main__':
	try:
		if not os.path.exists('Config.xlsx'):
			wb = Workbook()
			ws = wb.active
			ws.append(['Library', 'Loop Name', 'Start Page'])
			ws.column_dimensions['A'].width = 50.0
			ws.column_dimensions['B'].width = 50.0
			wb.save('Config.xlsx')
			input('Config.xlsx is generated.')
		else:
			wb = load_workbook('Config.xlsx')
			ws = wb.active
			table = list(ws.rows)

			Config(table)

			input('Finish!\nArguments file is generated.')
	except Exception as err:
		print('Exception: ' + repr(err))
		input(traceback.format_exc())
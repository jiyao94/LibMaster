##########################################################################################
'''
AUTHOR:			Yao Ji (jiyao94@126.com)
CREATED DATE:	2018/6/22
LAST UPDATE:	2018/7/19
DESCRIPTION:	This tool is used to read the I/O ports and the function blocks of the DPU
				configuration file and export to an Excel file. All the external ports and
				parametric function blocks should follow standard description form. The
				input can be a library file or a directory contains libraries. The output
				file will be exported to './Library' directory. This tool will also copy 
				the config files to this directory.
'''
##########################################################################################

import os, glob, readline, shutil, traceback
from collections import namedtuple
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, colors

#Argument tuple to store block information
Args = namedtuple('Args', 'blockName index desc inPointDir')

#function for auto-copmplete path input
def complete(text, state):
	return glob.glob(text + '*')[state]

#check whether the description contains the correct index format and description
def ContainIndex(str):
	sl = str.split(' ')
	if len(sl) < 2:
		return False
	elif len(sl[0]) != 5 or len(sl[1]) < 1:
		return False
	elif sl[0][0:2] in ['AI','AO','DI','DO','FB'] and sl[0][2:].isdigit():
		return True
	else:
		return False

#go through the list, find the index, add description
def AddDescription(lst, index, desc):
	for pin in lst:
		if pin.inPointDir and pin.index == index:
			lst.append(Args(pin.blockName, pin.index, desc, True))
			lst.remove(pin)
			break
		else:
			pass

#remove the path from file name
def RemovePath(fileName):
	if '/' in fileName:
		fileName = fileName.split('/')[-1]
	elif '\\' in fileName:
		fileName = fileName.split('\\')[-1]
	else:
		pass
	return fileName

#read through the file and store all the I/Os and FBs in lists
def ReadFile(fd):
	iList, oList, pList, pageList = [], [], [], []				#create lists to store inputs, outputs and function blocks
	it = iter(fd.readlines())
	pageBlock = False											#create 'page' flag, to read blocks inside pages
	pointInfo = False											#create 'point' flag, to read point config for blocks
	for line in it:
		if line.find('Page,') > -1:								#identify 'Page' block
			pageBlock = True									#set 'page' flag
			pageNo = line.split(', ')[1].split(':')[0]			#store page number
			pageList.append(pageNo)
		elif line.find('PageEnd') > -1:							#'Page' block ends
			pageBlock = False									#clear flag
		elif line.find('[POINT_DIR INFO]') > -1:
			pointInfo = True
		else:
			pass
		if pageBlock and line.find('Func,') > -1:				#identify 'Func' block
			words = line.split(', ')							#words contain function block information
			in_line, para_line, out_line = '', '', ''
			while True:
				next_line = next(it)
				if next_line.find('In=') > -1:					#next line is "In="
					in_line = next_line
				elif next_line.find('Para=') > -1:				#next line is "Para="
					para_line = next_line
				elif next_line.find('Out=') > -1:				#next line is "Out="
					out_line = next_line
				else:
					break
			if out_line.find('\\') > -1:						#check whether "Out=" contains description
				desc = out_line.split('\\')[1].split(',')[0]
			elif in_line.find('\\') > -1:						#check whether "In=" contains description
				desc = in_line.split('\\')[1].split(',')[0]
			else:												#otherwise, pin has no description, it is not an external I/O
				desc = 'NO DESCRIPTIOIN'
				pass
			pointArg = None
			if ContainIndex(desc):								#check whether 'desc' contains the correct index and description
				pointArg = Args(words[1], desc[0:5], desc[6:], False)
			else:
				if len(para_line.split(',')) > 1:				#check whether "Para="contains index
					desc = para_line.split(',')[1]				#if it contains correct index but no description
					if len(desc) >= 5 and desc[0:2] in ['AI','AO','DI','DO','FB'] and desc[2:5].isdigit():
						pointArg = Args(words[1], desc[0:5], '', True)
					else:
						pass
				else:
					pass
			if pointArg is not None:
				if pointArg.index[1] == 'I':					#indicate input port
					iList.append(pointArg)
				elif pointArg.index[1] == 'O':					#indicate output port
					oList.append(pointArg)
				elif pointArg.index[0:2] == 'FB':				#indicate Function Block with parameter option
					pList.append(pointArg)
				else:
					pass
		elif pointInfo and line.find('=') > -1 and line.find(',') > -1:		#indentify point configs
			word = line.split('=')[0]							#if tag has index form, find index from lists and add description
			if len(word) >= 5 and word[0:2] in ['AI','AO','DI','DO','FB'] and word[2:5].isdigit():
				if word[1] == 'I':
					AddDescription(iList, word, line.split(',')[1])
				elif word[1] == 'O':
					AddDescription(oList, word, line.split(',')[1])
				elif word[0:2] == 'FB':
					AddDescription(pList, word, line.split(',')[1])
			else:
				pass
		else:
			pass
	return iList, oList, pList, pageList

#helper function for WriteExcel()
def StyleMerge(ws, val, range):
	ws.merge_cells(range)
	cell = Cell(ws, value=val)
	cell.font = Font(bold=True, color=colors.WHITE)
	cell.fill = PatternFill(fill_type='solid', fgColor=colors.BLACK)
	cell.alignment = Alignment(horizontal='center')
	return cell
def StyleRange(ws, str_list, isBold, fillColor):
	row_list = []
	for str_i in str_list:
		cell = Cell(ws, value=str_i)
		cell.font = Font(bold=isBold)
		cell.fill = PatternFill(fill_type='solid', fgColor=fillColor)
		row_list.append(cell)
	return row_list
def StyleBorder(ws, start_row, end_row, columns):
	thin = Side(border_style='thin', color=colors.BLACK)
	for i in range(start_row, end_row):
		for j in range(columns):
			cell = ws.cell(row = i + 1, column = j + 1)
			cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

#write the I/O list into Excel
#openpyxl is used to write .xlsx format
def WriteExcel(fileName, iList, oList, pList, pageList):
	wb = Workbook()
	ws = wb.active

	#first write fileName as title
	ws.merge_cells('A1:F1')
	cell = Cell(ws, value=fileName.split('.')[0])
	cell.font = Font(bold=True)
	cell.alignment = Alignment(horizontal='center')
	ws.append([cell])

	ws.append([None])

	#write input ports list
	ws.append([StyleMerge(ws, 'INPUTS', 'A3:C3')])
	ws.append(StyleRange(ws, ['INDEX', 'POINT DESCRIPTION', 'TAG NAME'], True, colors.YELLOW))
	for pin in iList:
		ws.append([pin.index, pin.desc])
	StyleBorder(ws, 3, len(iList) + 4, 3)

	ws.append([None])

	#write output ports list
	ws.append([StyleMerge(ws, 'OUTPUTS', 'A{0}:D{0}'.format(len(iList) + 6))])
	ws.append(StyleRange(ws, ['INDEX', 'FUNCTION BLOCK DESCRIPTION', 'TAG NAME', 'DESCRIPTION'], True, colors.YELLOW))
	for pin in oList:
		if not pin.inPointDir:
			ws.append(StyleRange(ws, [pin.index, pin.desc], False, colors.GREEN))
		else:
			ws.append([pin.index, pin.desc])
	StyleBorder(ws, len(iList) + 6, len(iList + oList) + 7, 4)

	ws.append([None])

	#write function blocks list
	ws.append([StyleMerge(ws, 'PARAMETERS', 'A{0}:F{0}'.format(len(iList + oList) + 9))])
	ws.append(StyleRange(ws, ['INDEX', 'POINT DESCRIPTION', 'TAG NAME', 'DESCRIPTION', 'PARA', 'VALUE'], True, colors.YELLOW))
	for pin in pList:
		if not pin.inPointDir:
			ws.append(StyleRange(ws, [pin.index, pin.desc], False, colors.GREEN))
		else:
			ws.append([pin.index, pin.desc])
	StyleBorder(ws, len(iList + oList) + 9, len(iList + oList + pList) + 10, 6)

	#change the width of the columns
	ws.column_dimensions['B'].width = 30.0
	ws.column_dimensions['C'].width = 15.0
	ws.column_dimensions['D'].width = 30.0

	#store page information in seperate sheet
	ws = wb.create_sheet('Info')
	ws.append(['Page count', len(pageList)])
	ws.append(['Page list'] + pageList)

	#file is outputed under './Library' directory
	wb.save('Library/' + RemovePath(fileName).replace('.txt', '.xlsx'))

#main import function
def Import(dirs, OutputFunc = print):
	files = []
	for fileName in dirs:
		#start reading file
		fd = open(fileName, encoding='utf-16')
		OutputFunc(RemovePath(fileName) + ':\n\tReading... ')
		iList, oList, pList, pageList = ReadFile(fd)
		fd.close()

		#Option: sort the list
		iList.sort(key= lambda ele: ele.index)
		oList.sort(key= lambda ele: ele.index)
		pList.sort(key= lambda ele: ele.index)

		#write the lists into Excel
		OutputFunc('\tWriting... ')
		WriteExcel(RemovePath(fileName), iList, oList, pList, pageList)
		OutputFunc('\tDone!')

		#copy file to library
		try:
			shutil.copy2(fileName, 'Library/' + RemovePath(fileName))
			files.append(RemovePath(fileName))
		except shutil.SameFileError: 
			pass
	return files

#start main
if __name__ == '__main__':
	try:
		#create './Library' directory if not exist
		if not os.path.exists('Library'):
			os.mkdir('Library')

		#open selected file or directory
		readline.set_completer(complete)					#add path auto-complete
		readline.parse_and_bind("tab: complete")			#using tab for auto-complete
		readline.set_completer_delims('\t')					#just like in shell

		dirName = input('Import which file or directory: ')
		if os.path.isdir(dirName):							#input is a directory
			dirs = os.listdir(dirName)						#list all file names under directory
			for i in range(len(dirs) - 1, -1, -1):			#discard invalid files
				if dirs[i].find('.txt') < 0:				#
					dirs.pop(i)								#
				else:										#for valid files, we need to
					dirs[i] = dirName + '/' + dirs[i]		#add the path before file names
		else:												#input is a file,
			dirs = [dirName]								#just create a single element list

		Import(dirs)

		#finish
		if len(dirs) < 1:
			input('Empty directory or no valid file in the directory!')
		else:
			input('Finish!\nI/O Lists are exported under \'./Library\' directory.')

	#error report
	except Exception as err:
		print('\nException: ' + repr(err))
		input(traceback.format_exc())